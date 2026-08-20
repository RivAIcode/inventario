import os
import hashlib
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, jsonify, make_response, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from PIL import Image
import pandas as pd
from io import BytesIO
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import base64
import shutil
import json
import pytz

app = Flask(__name__)
app.secret_key = 'tu_clave_secreta_aqui_cambiala_por_una_segura'

def get_chile_time():
    chile_tz = pytz.timezone('America/Santiago')
    return datetime.now(chile_tz)

EMAIL_CONFIG = {
    "activado": True,
    "remitente": "inventarioequipos55@gmail.com",
    "password": "jqah iuho hjuo bpbu",
    "destinatarios": ["driveram@suralis.cl"],
}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(BASE_DIR, "inventario.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'media')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.config['MAX_IMAGENES'] = 3
app.config['MAX_DOCUMENTOS'] = 3

for folder in ['equipos_fotos', 'equipos_docs', 'reportes']:
    os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], folder), exist_ok=True)

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor, inicie sesión para acceder'

LOCALIDADES = [
    "Panguipulli", "Lanco", "San Jose", "Mafil", "Los Lagos", "Corral",
    "Paillaco", "Futrono", "Lago Ranco", "Rio Bueno", "La Union", "Cocule",
    "Biofiltro", "San Pablo", "Osorno", "Rio Negro", "Purranque", "Corte Alto",
    "Frutillar", "Llanquihue", "Fresia", "Puerto Varas", "Los Muermos", "Alerce",
    "Maullin", "Calbuco", "Puerto Montt", "Ancud", "Dalcahue", "Castro", "Chonchi",
    "Quellon", "Achao", "Futaleufu", "Chaiten", "Calidad"
]

TIPOS_EQUIPO = [
    "Colorimetro DR300", "Pocket", "Multiparametro",
    "Muestreador afluente", "Muestreador efluente",
    "pHmetro lodo", "pHmetro agua",
    "Gelex cloro", "Stabcal", "Termometro patron AS"
]

SUBCATEGORIAS = ["cloro", "hierro", "manganeso", "fluor"]
ESTADOS = ["Operativo", "Volante", "Mantencion", "Contrastacion", "Prestado", "Fuera de Servicio"]
AREAS = ["AP", "AS"]

TIPOS_CON_PHMETRO = ["pHmetro lodo", "pHmetro agua"]
TIPOS_CON_CERTIFICADO = ["Gelex cloro", "Stabcal", "Termometro patron AS"]
TIPOS_CON_VENCIMIENTO_MANUAL = ["Gelex cloro", "Stabcal"]
TIPOS_CON_VENCIMIENTO_AUTOMATICO = ["Termometro patron AS"]
TIPOS_CON_CONTRASTACION = ["Colorimetro DR300", "Pocket", "Multiparametro"]

class Usuario(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    nombre = db.Column(db.String(100), nullable=False)
    es_admin = db.Column(db.Boolean, default=False)
    cambiar_password = db.Column(db.Boolean, default=True)
    fecha_registro = db.Column(db.DateTime, default=get_chile_time)

class ConfiguracionCorreo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    frecuencia = db.Column(db.String(20), default='diario')
    hora_envio = db.Column(db.String(5), default='08:00')
    destinatarios = db.Column(db.Text, default='')
    activo = db.Column(db.Boolean, default=True)
    ultimo_envio = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=get_chile_time)
    updated_at = db.Column(db.DateTime, default=get_chile_time, onupdate=get_chile_time)

class Responsable(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), unique=True, nullable=False)
    email = db.Column(db.String(100), nullable=True)
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=get_chile_time)

class Equipo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nro_int = db.Column(db.String(50), unique=True, nullable=False)
    nro_serie = db.Column(db.String(100))
    area = db.Column(db.String(10))
    localidad = db.Column(db.String(100))
    responsable = db.Column(db.String(100))
    tipo_equipo = db.Column(db.String(50))
    subcategoria = db.Column(db.String(50))
    modelo_sonda = db.Column(db.String(100))
    modelo_equipo = db.Column(db.String(100))
    nro_serie_sonda = db.Column(db.String(100))
    estado = db.Column(db.String(50), default='Operativo')
    servicio_tecnico = db.Column(db.String(200))
    fecha_envio_laboratorio = db.Column(db.Date, nullable=True)
    fecha_contrastacion = db.Column(db.Date)
    fecha_certificado = db.Column(db.Date)
    fecha_vencimiento_insumo = db.Column(db.Date)
    fecha_ultima_mantencion = db.Column(db.Date)
    fecha_retorno_mantencion = db.Column(db.Date)
    observaciones = db.Column(db.Text)
    fecha_creacion = db.Column(db.DateTime, default=get_chile_time)
    ultima_actualizacion = db.Column(db.DateTime, default=get_chile_time, onupdate=get_chile_time)
    version = db.Column(db.Integer, default=1)

    # ====== NUEVOS CAMPOS ======
    # Contrastación
    fecha_certificado_contrastacion = db.Column(db.Date, nullable=True)
    nro_informe_contrastacion = db.Column(db.String(100), nullable=True)

    # Mantención
    fecha_certificado_mantencion = db.Column(db.Date, nullable=True)
    nro_informe_mantencion = db.Column(db.String(100), nullable=True)

    # Despacho general
    fecha_despacho = db.Column(db.Date, nullable=True)

    # Termómetros e Insumos
    nro_certificado_termometro = db.Column(db.String(100), nullable=True)
    fecha_contrastacion_termometro = db.Column(db.Date, nullable=True)

class Archivo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    equipo_id = db.Column(db.Integer, db.ForeignKey('equipo.id'), nullable=False)
    tipo = db.Column(db.String(20))
    nombre_original = db.Column(db.String(200))
    nombre_archivo = db.Column(db.String(200), unique=True)
    fecha_subida = db.Column(db.DateTime, default=get_chile_time)
    equipo = db.relationship('Equipo', backref=db.backref('archivos', cascade='all, delete-orphan'))

class Historial(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    equipo_id = db.Column(db.Integer, db.ForeignKey('equipo.id'), nullable=False)
    accion = db.Column(db.String(50))
    detalle = db.Column(db.Text)
    usuario = db.Column(db.String(80))
    responsable = db.Column(db.String(100))
    fecha_hora = db.Column(db.DateTime, default=get_chile_time)
    equipo = db.relationship('Equipo', backref=db.backref('historial', cascade='all, delete-orphan'))

class BloqueoEquipo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    equipo_id = db.Column(db.Integer, db.ForeignKey('equipo.id'), unique=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    usuario_nombre = db.Column(db.String(100))
    fecha_bloqueo = db.Column(db.DateTime, default=get_chile_time)
    expira = db.Column(db.DateTime)
    equipo = db.relationship('Equipo', backref=db.backref('bloqueo', uselist=False))
    usuario = db.relationship('Usuario', backref=db.backref('bloqueos'))

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

def formatear_fecha(fecha):
    if fecha:
        return fecha.strftime('%d/%m/%Y')
    return ''

def formatear_fecha_hora(fecha):
    if fecha:
        return fecha.strftime('%d/%m/%Y %H:%M')
    return ''

def calcular_vencimiento_contrastacion(fecha_contrastacion):
    if fecha_contrastacion:
        return fecha_contrastacion + timedelta(days=180)
    return None

def calcular_vencimiento_insumo(fecha_certificado):
    if fecha_certificado:
        return fecha_certificado + timedelta(days=365)
    return None

def dias_habiles(fecha_inicio, fecha_fin):
    if not fecha_inicio or not fecha_fin:
        return 0
    dias = 0
    current = fecha_inicio
    while current <= fecha_fin:
        if current.weekday() < 5:
            dias += 1
        current += timedelta(days=1)
    return dias

def limpiar_bloqueos_expirados():
    from datetime import datetime
    ahora = datetime.now()
    bloqueos = BloqueoEquipo.query.all()
    for bloqueo in bloqueos:
        if bloqueo.expira and bloqueo.expira.replace(tzinfo=None) < ahora:
            db.session.delete(bloqueo)
    db.session.commit()

def bloquear_equipo(equipo_id, usuario_id, usuario_nombre):
    from datetime import datetime, timedelta
    limpiar_bloqueos_expirados()
    ahora = datetime.now()
    bloqueo = BloqueoEquipo.query.filter_by(equipo_id=equipo_id).first()
    if bloqueo:
        if bloqueo.fecha_bloqueo + timedelta(seconds=10) > ahora:
            return False, bloqueo.usuario_nombre
        else:
            db.session.delete(bloqueo)
            db.session.commit()
    nuevo_bloqueo = BloqueoEquipo(
        equipo_id=equipo_id,
        usuario_id=usuario_id,
        usuario_nombre=usuario_nombre,
        fecha_bloqueo=ahora,
        expira=ahora + timedelta(seconds=10)
    )
    db.session.add(nuevo_bloqueo)
    db.session.commit()
    return True, None

def liberar_bloqueo(equipo_id):
    bloqueo = BloqueoEquipo.query.filter_by(equipo_id=equipo_id).first()
    if bloqueo:
        db.session.delete(bloqueo)
        db.session.commit()
    return True

def verificar_bloqueo(equipo_id):
    limpiar_bloqueos_expirados()
    bloqueo = BloqueoEquipo.query.filter_by(equipo_id=equipo_id).first()
    if bloqueo:
        return True, bloqueo.usuario_nombre, bloqueo.fecha_bloqueo
    return False, None, None

def obtener_alerta(equipo):
    """Retorna un diccionario con la información de alerta para un equipo"""
    tipo = equipo.tipo_equipo
    hoy = get_chile_time().date()

    # ====== PRIORIDAD 1: EQUIPO EN CONTRASTACIÓN ======
    if equipo.estado == 'Contrastacion' and equipo.fecha_envio_laboratorio:
        dias = (hoy - equipo.fecha_envio_laboratorio).days
        return {
            'texto': f'📊 En contrastación ({dias} días)',
            'clase': 'alerta-contrastacion',
            'icono': '📊',
            'dias_texto': f'{dias}d'
        }

    # ====== PRIORIDAD 1: EQUIPO EN MANTENCIÓN ======
    if equipo.estado == 'Mantencion' and equipo.fecha_ultima_mantencion:
        dias = (hoy - equipo.fecha_ultima_mantencion).days
        return {
            'texto': f'🔧 En mantención ({dias} días)',
            'clase': 'alerta-mantencion',
            'icono': '🔧',
            'dias_texto': f'{dias}d'
        }

    # ====== PRIORIDAD 2: TERMÓMETRO PATRÓN - CONTRASTACIÓN (5 días) ======
    if tipo == 'Termometro patron AS' and equipo.fecha_contrastacion_termometro:
        fecha_venc = equipo.fecha_contrastacion_termometro + timedelta(days=30)
        dias = (fecha_venc - hoy).days
        if dias < 0:
            return {
                'texto': f'⚠️ VENCIDA contrastación termómetro (venció hace {abs(dias)} días)',
                'clase': 'alerta-vencido',
                'icono': '🔴',
                'dias_texto': f'{abs(dias)}d'
            }
        elif dias <= 5:
            return {
                'texto': f'⚠️ Vence contrastación termómetro en {dias} días',
                'clase': 'alerta-proximo',
                'icono': '🟡',
                'dias_texto': f'{dias}d'
            }

    # ====== PRIORIDAD 2: TERMÓMETRO PATRÓN - CERTIFICADO (30 días) ======
    if tipo == 'Termometro patron AS' and equipo.fecha_certificado:
        fecha_venc = calcular_vencimiento_insumo(equipo.fecha_certificado)
        if fecha_venc:
            dias = (fecha_venc - hoy).days
            if dias < 0:
                return {
                    'texto': f'⚠️ VENCIDO certificado termómetro (venció hace {abs(dias)} días)',
                    'clase': 'alerta-vencido',
                    'icono': '🔴',
                    'dias_texto': f'{abs(dias)}d'
                }
            elif dias <= 30:
                return {
                    'texto': f'⚠️ Vence certificado termómetro en {dias} días',
                    'clase': 'alerta-proximo',
                    'icono': '🟡',
                    'dias_texto': f'{dias}d'
                }
            return {
                'texto': f'Certificado termómetro: {formatear_fecha(fecha_venc)}',
                'clase': '',
                'icono': '✅',
                'dias_texto': ''
            }

    # ====== PRIORIDAD 2: EQUIPOS CON CONTRASTACIÓN (Colorímetros, Pocket, Multiparamétro) ======
    if tipo in TIPOS_CON_CONTRASTACION and equipo.fecha_contrastacion:
        fecha_venc = calcular_vencimiento_contrastacion(equipo.fecha_contrastacion)
        if fecha_venc:
            dias = (fecha_venc - hoy).days
            if dias < 0:
                return {
                    'texto': f'⚠️ VENCIDO (venció: {formatear_fecha(fecha_venc)})',
                    'clase': 'alerta-vencido',
                    'icono': '🔴',
                    'dias_texto': f'{abs(dias)}d'
                }
            elif dias <= 15:
                return {
                    'texto': f'⚠️ VENCE EN {dias} DÍAS ({formatear_fecha(fecha_venc)})',
                    'clase': 'alerta-proximo',
                    'icono': '🟡',
                    'dias_texto': f'{dias}d'
                }
            return {
                'texto': f'Vence: {formatear_fecha(fecha_venc)}',
                'clase': '',
                'icono': '✅',
                'dias_texto': ''
            }

    # ====== PRIORIDAD 2: EQUIPOS CON CERTIFICADO (Gelex, Stabcal) ======
    if tipo in TIPOS_CON_CERTIFICADO and tipo != 'Termometro patron AS':
        fecha_venc = equipo.fecha_vencimiento_insumo
        if fecha_venc:
            dias = (fecha_venc - hoy).days
            if dias < 0:
                return {
                    'texto': f'⚠️ VENCIDO (venció: {formatear_fecha(fecha_venc)})',
                    'clase': 'alerta-vencido',
                    'icono': '🔴',
                    'dias_texto': f'{abs(dias)}d'
                }
            elif dias <= 15:
                return {
                    'texto': f'⚠️ VENCE EN {dias} DÍAS ({formatear_fecha(fecha_venc)})',
                    'clase': 'alerta-proximo',
                    'icono': '🟡',
                    'dias_texto': f'{dias}d'
                }
            return {
                'texto': f'Vence: {formatear_fecha(fecha_venc)}',
                'clase': '',
                'icono': '✅',
                'dias_texto': ''
            }

    # Sin alerta
    return {
        'texto': '',
        'clase': '',
        'icono': '➖',
        'dias_texto': ''
    }

def archivo_permitido(filename):
    extensiones = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in extensiones

def comprimir_imagen(ruta, calidad=70):
    try:
        img = Image.open(ruta)
        img.thumbnail((800, 800))
        img.save(ruta, optimize=True, quality=calidad)
    except:
        pass

def calcular_tiempo_inactividad(equipo):
    if equipo.fecha_ultima_mantencion:
        entrada = equipo.fecha_ultima_mantencion
        if equipo.fecha_retorno_mantencion:
            return (equipo.fecha_retorno_mantencion - entrada).days
        else:
            return (get_chile_time().date() - entrada).days
    return 0

def enviar_correo_alertas_general(vencidos, proximos, destinatarios):
    if not destinatarios:
        return False
    if not vencidos and not proximos:
        return False
    try:
        import yagmail
    except ImportError:
        return False
    try:
        ahora = get_chile_time()
        asunto = f"📊 ALERTA GENERAL - Equipos con vencimiento - {ahora.strftime('%d/%m/%Y')}"

        html = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="font-family: Arial, sans-serif;">
<div style="max-width:800px; margin:0 auto; background:white; padding:20px; border-radius:12px;">
<h2 style="color:#1E88E5;">📊 SISTEMA DE INVENTARIO - ALERTA GENERAL</h2>
<p>Se detectaron equipos que requieren atención:</p>
"""
        if vencidos:
            html += '<h3 style="color:#c62828;">⚠️ EQUIPOS VENCIDOS</h3>'
            html += '<table border="1" cellpadding="8" style="border-collapse:collapse; width:100%;">'
            html += '<tr style="background:#1E88E5; color:white;"><th>Nro Int</th><th>Nro Serie</th><th>Equipo</th><th>Localidad</th><th>Responsable</th><th>Alerta</th></tr>'
            for e in vencidos:
                alerta = obtener_alerta(e)
                html += f'<tr><td style="padding:8px;">{e.nro_int}</td><td style="padding:8px;">{e.nro_serie or "-"}</td><td style="padding:8px;">{e.tipo_equipo or "-"}</td><td style="padding:8px;">{e.localidad or "-"}</td><td style="padding:8px;">{e.responsable or "-"}</td><td style="padding:8px; color:#c62828;">{alerta["texto"]}</td></tr>'
            html += '</table><br>'

        if proximos:
            html += '<h3 style="color:#f57c00;">⚠️ EQUIPOS PRÓXIMOS A VENCER</h3>'
            html += '<table border="1" cellpadding="8" style="border-collapse:collapse; width:100%;">'
            html += '<tr style="background:#1E88E5; color:white;"><th>Nro Int</th><th>Nro Serie</th><th>Equipo</th><th>Localidad</th><th>Responsable</th><th>Alerta</th></tr>'
            for e in proximos:
                alerta = obtener_alerta(e)
                html += f'<tr><td style="padding:8px;">{e.nro_int}</td><td style="padding:8px;">{e.nro_serie or "-"}</td><td style="padding:8px;">{e.tipo_equipo or "-"}</td><td style="padding:8px;">{e.localidad or "-"}</td><td style="padding:8px;">{e.responsable or "-"}</td><td style="padding:8px; color:#f57c00;">{alerta["texto"]}</td></tr>'
            html += '</table><br>'

        html += """
<p>Por favor, regulariza la situación de estos equipos.</p>
<hr>
<p style="font-size:11px; color:#999;">Mensaje automático - Sistema de Inventario</p>
</div>
</body>
</html>"""

        yag = yagmail.SMTP(EMAIL_CONFIG["remitente"], EMAIL_CONFIG["password"])
        yag.send(to=destinatarios, subject=asunto, contents=html)
        print(f"✅ Correo general enviado a: {', '.join(destinatarios)}")
        return True
    except Exception as e:
        print(f"❌ Error en correo general: {e}")
        return False

def enviar_correo_alertas_responsable(email_responsable, nombre_responsable, vencidos, proximos):
    if not email_responsable:
        return False
    if not vencidos and not proximos:
        return False
    try:
        import yagmail
    except ImportError:
        return False
    try:
        ahora = get_chile_time()
        asunto = f"📊 ALERTA - Equipos a tu cargo - {ahora.strftime('%d/%m/%Y')}"

        html = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="font-family: Arial, sans-serif;">
<div style="max-width:600px; margin:0 auto; background:white; padding:20px; border-radius:12px;">
<h2 style="color:#1E88E5;">📊 SISTEMA DE INVENTARIO</h2>
<p>Hola <strong>{nombre_responsable}</strong>,</p>
<p>Estos son los equipos que tienes a tu cargo y requieren atención:</p>
"""
        if vencidos:
            html += '<h3 style="color:#c62828;">⚠️ EQUIPOS VENCIDOS</h3>'
            html += '<table border="1" cellpadding="8" style="border-collapse:collapse; width:100%;">'
            html += '<tr style="background:#1E88E5; color:white;"><th>Nro Int</th><th>Nro Serie</th><th>Equipo</th><th>Localidad</th><th>Alerta</th></tr>'
            for e in vencidos:
                alerta = obtener_alerta(e)
                html += f'<tr><td style="padding:8px;">{e.nro_int}</td><td style="padding:8px;">{e.nro_serie or "-"}</td><td style="padding:8px;">{e.tipo_equipo or "-"}</td><td style="padding:8px;">{e.localidad or "-"}</td><td style="padding:8px; color:#c62828;">{alerta["texto"]}</td></tr>'
            html += '</table><br>'

        if proximos:
            html += '<h3 style="color:#f57c00;">⚠️ EQUIPOS PRÓXIMOS A VENCER</h3>'
            html += '<table border="1" cellpadding="8" style="border-collapse:collapse; width:100%;">'
            html += '<tr style="background:#1E88E5; color:white;"><th>Nro Int</th><th>Nro Serie</th><th>Equipo</th><th>Localidad</th><th>Alerta</th></tr>'
            for e in proximos:
                alerta = obtener_alerta(e)
                html += f'<tr><td style="padding:8px;">{e.nro_int}</td><td style="padding:8px;">{e.nro_serie or "-"}</td><td style="padding:8px;">{e.tipo_equipo or "-"}</td><td style="padding:8px;">{e.localidad or "-"}</td><td style="padding:8px; color:#f57c00;">{alerta["texto"]}</td></tr>'
            html += '</table><br>'

        html += """
<p>Por favor, regulariza la situación de estos equipos.</p>
<hr>
<p style="font-size:11px; color:#999;">Mensaje automático - Sistema de Inventario</p>
</div>
</body>
</html>"""

        yag = yagmail.SMTP(EMAIL_CONFIG["remitente"], EMAIL_CONFIG["password"])
        yag.send(to=email_responsable, subject=asunto, contents=html)
        print(f"✅ Correo enviado a {nombre_responsable} ({email_responsable})")
        return True
    except Exception as e:
        print(f"❌ Error al enviar a {email_responsable}: {e}")
        return False

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    session.pop('_flashes', None)
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form['username']
        password = hashlib.sha256(request.form['password'].encode()).hexdigest()
        usuario = Usuario.query.filter_by(username=username, password=password).first()
        if usuario:
            login_user(usuario)
            verificar_y_crear_respaldo()
            if usuario.cambiar_password:
                flash('Debes cambiar tu contraseña antes de continuar', 'warning')
                return redirect(url_for('cambiar_password'))
            return redirect(url_for('dashboard'))
        flash('Usuario o contraseña incorrectos', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    equipos = Equipo.query.order_by(Equipo.nro_int).all()
    total = len(equipos)
    operativo = sum(1 for e in equipos if e.estado == 'Operativo')
    mantencion = sum(1 for e in equipos if e.estado == 'Mantencion')
    contrastacion = sum(1 for e in equipos if e.estado == 'Contrastacion')
    prestado = sum(1 for e in equipos if e.estado == 'Prestado')
    volante = sum(1 for e in equipos if e.estado == 'Volante')
    fuera_servicio = sum(1 for e in equipos if e.estado == 'Fuera de Servicio')
    vencidos = 0
    proximos = 0
    for e in equipos:
        alerta = obtener_alerta(e)
        if 'VENCIDO' in alerta['texto'] or 'VENCIDA' in alerta['texto']:
            vencidos += 1
        elif 'VENCE' in alerta['texto']:
            proximos += 1

    responsables_unicos = Responsable.query.order_by(Responsable.nombre).all()

    return render_template('dashboard.html',
                         equipos=equipos,
                         total=total,
                         operativo=operativo,
                         mantencion=mantencion,
                         contrastacion=contrastacion,
                         prestado=prestado,
                         volante=volante,
                         fuera_servicio=fuera_servicio,
                         vencidos=vencidos,
                         proximos=proximos,
                         formatear_fecha=formatear_fecha,
                         obtener_alerta=obtener_alerta,
                         LOCALIDADES=LOCALIDADES,
                         TIPOS_EQUIPO=TIPOS_EQUIPO,
                         SUBCATEGORIAS=SUBCATEGORIAS,
                         ESTADOS=ESTADOS,
                         AREAS=AREAS,
                         TIPOS_CON_PHMETRO=TIPOS_CON_PHMETRO,
                         TIPOS_CON_CERTIFICADO=TIPOS_CON_CERTIFICADO,
                         TIPOS_CON_CONTRASTACION=TIPOS_CON_CONTRASTACION,
                         TIPOS_CON_VENCIMIENTO_MANUAL=TIPOS_CON_VENCIMIENTO_MANUAL,
                         TIPOS_CON_VENCIMIENTO_AUTOMATICO=TIPOS_CON_VENCIMIENTO_AUTOMATICO,
                         now=get_chile_time(),
                         verificar_bloqueo=verificar_bloqueo,
                         responsables_unicos=responsables_unicos)

@app.route('/equipo/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_equipo():
    if request.method == 'POST':
        try:
            equipo = Equipo(
                nro_int=request.form['nro_int'],
                nro_serie=request.form.get('nro_serie', ''),
                area=request.form.get('area', ''),
                localidad=request.form.get('localidad', ''),
                responsable=request.form.get('responsable', ''),
                tipo_equipo=request.form.get('tipo_equipo', ''),
                subcategoria=request.form.get('subcategoria', ''),
                modelo_sonda=request.form.get('modelo_sonda', ''),
                modelo_equipo=request.form.get('modelo_equipo', ''),
                nro_serie_sonda=request.form.get('nro_serie_sonda', ''),
                estado=request.form.get('estado', 'Operativo'),
                servicio_tecnico=request.form.get('servicio_tecnico', ''),
                observaciones=request.form.get('observaciones', ''),
                # Nuevos campos
                nro_informe_contrastacion=request.form.get('nro_informe_contrastacion', ''),
                nro_informe_mantencion=request.form.get('nro_informe_mantencion', ''),
                nro_certificado_termometro=request.form.get('nro_certificado_termometro', ''),
                fecha_certificado_contrastacion=datetime.strptime(request.form['fecha_certificado_contrastacion'], '%Y-%m-%d').date() if request.form.get('fecha_certificado_contrastacion') else None,
                fecha_certificado_mantencion=datetime.strptime(request.form['fecha_certificado_mantencion'], '%Y-%m-%d').date() if request.form.get('fecha_certificado_mantencion') else None,
                fecha_contrastacion_termometro=datetime.strptime(request.form['fecha_contrastacion_termometro'], '%Y-%m-%d').date() if request.form.get('fecha_contrastacion_termometro') else None,
                fecha_despacho=datetime.strptime(request.form['fecha_despacho'], '%Y-%m-%d').date() if request.form.get('fecha_despacho') else None
            )
            if request.form.get('fecha_contrastacion'):
                equipo.fecha_contrastacion = datetime.strptime(request.form['fecha_contrastacion'], '%Y-%m-%d').date()
            if request.form.get('fecha_certificado'):
                equipo.fecha_certificado = datetime.strptime(request.form['fecha_certificado'], '%Y-%m-%d').date()
            if request.form.get('fecha_vencimiento_insumo'):
                equipo.fecha_vencimiento_insumo = datetime.strptime(request.form['fecha_vencimiento_insumo'], '%Y-%m-%d').date()
            if request.form.get('fecha_ultima_mantencion'):
                equipo.fecha_ultima_mantencion = datetime.strptime(request.form['fecha_ultima_mantencion'], '%Y-%m-%d').date()
            if request.form.get('fecha_retorno_mantencion'):
                equipo.fecha_retorno_mantencion = datetime.strptime(request.form['fecha_retorno_mantencion'], '%Y-%m-%d').date()
            if request.form.get('fecha_envio_laboratorio'):
                equipo.fecha_envio_laboratorio = datetime.strptime(request.form['fecha_envio_laboratorio'], '%Y-%m-%d').date()
            if equipo.tipo_equipo in TIPOS_CON_VENCIMIENTO_AUTOMATICO and equipo.fecha_certificado:
                equipo.fecha_vencimiento_insumo = calcular_vencimiento_insumo(equipo.fecha_certificado)
            db.session.add(equipo)
            db.session.flush()
            if equipo.observaciones:
                historial_obs = Historial(
                    equipo_id=equipo.id, accion='OBSERVACION',
                    detalle=equipo.observaciones,
                    usuario=current_user.username, responsable=equipo.responsable
                )
                db.session.add(historial_obs)
            historial = Historial(
                equipo_id=equipo.id, accion='CREACION',
                detalle=f'Equipo creado: {equipo.nro_int}',
                usuario=current_user.username, responsable=equipo.responsable
            )
            db.session.add(historial)
            archivos = request.files.getlist('archivos')
            for archivo in archivos:
                if archivo and archivo.filename and archivo_permitido(archivo.filename):
                    ext = archivo.filename.rsplit('.', 1)[1].lower()
                    tipo = 'foto' if ext in {'png', 'jpg', 'jpeg', 'gif'} else 'documento'
                    nombre_seguro = f"{equipo.id}_{get_chile_time().strftime('%Y%m%d_%H%M%S')}_{secure_filename(archivo.filename)}"
                    subcarpeta = 'equipos_fotos' if tipo == 'foto' else 'equipos_docs'
                    ruta = os.path.join(app.config['UPLOAD_FOLDER'], subcarpeta, nombre_seguro)
                    archivo.save(ruta)
                    if tipo == 'foto':
                        comprimir_imagen(ruta)
                    archivo_db = Archivo(
                        equipo_id=equipo.id, tipo=tipo,
                        nombre_original=archivo.filename, nombre_archivo=nombre_seguro
                    )
                    db.session.add(archivo_db)
            db.session.commit()
            flash('Equipo agregado exitosamente', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    responsables_unicos = Responsable.query.order_by(Responsable.nombre).all()
    return render_template('equipo_form.html',
                         LOCALIDADES=LOCALIDADES, TIPOS_EQUIPO=TIPOS_EQUIPO,
                         SUBCATEGORIAS=SUBCATEGORIAS, ESTADOS=ESTADOS, AREAS=AREAS,
                         TIPOS_CON_PHMETRO=TIPOS_CON_PHMETRO,
                         TIPOS_CON_CERTIFICADO=TIPOS_CON_CERTIFICADO,
                         TIPOS_CON_VENCIMIENTO_MANUAL=TIPOS_CON_VENCIMIENTO_MANUAL,
                         TIPOS_CON_VENCIMIENTO_AUTOMATICO=TIPOS_CON_VENCIMIENTO_AUTOMATICO,
                         TIPOS_CON_CONTRASTACION=TIPOS_CON_CONTRASTACION,
                         responsables_unicos=responsables_unicos)

@app.route('/equipo/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_equipo(id):
    equipo = Equipo.query.get_or_404(id)

    # ====== LOGS DE DEPURACIÓN ======
    if request.method == 'POST':
        print("=" * 60)
        print("📥 EDITAR_EQUIPO - DATOS RECIBIDOS:")
        print(f"  ID del equipo: {id}")
        for key, value in request.form.items():
            print(f"  {key} = '{value}'")
        print("=" * 60)
        print(f"📌 estado recibido: '{request.form.get('estado', 'NO ENVIADO')}'")
        print("=" * 60)

    if request.method == 'GET':
        bloqueado, nombre_usuario, fecha_bloqueo = verificar_bloqueo(id)
        if bloqueado and nombre_usuario != current_user.nombre:
            flash(f'⚠️ El equipo está siendo editado por {nombre_usuario}. No puedes editarlo hasta que termine.', 'error')
            return redirect(url_for('dashboard'))
        exito, nombre_existente = bloquear_equipo(id, current_user.id, current_user.nombre)
        if not exito:
            flash(f'⚠️ El equipo está siendo editado por {nombre_existente}. Intenta más tarde.', 'error')
            return redirect(url_for('dashboard'))

    if request.method == 'POST':
        try:
            estado_anterior = equipo.estado
            responsable_anterior = equipo.responsable
            modelo_sonda_anterior = equipo.modelo_sonda
            modelo_equipo_anterior = equipo.modelo_equipo
            nro_serie_sonda_anterior = equipo.nro_serie_sonda
            observaciones_anteriores = equipo.observaciones
            fecha_contrastacion_anterior = equipo.fecha_contrastacion
            fecha_envio_laboratorio_anterior = equipo.fecha_envio_laboratorio
            fecha_ultima_mantencion_anterior = equipo.fecha_ultima_mantencion

            nuevas_observaciones = request.form.get('observaciones', '')

            # ====== ACTUALIZAR TODOS LOS CAMPOS ======
            equipo.nro_serie = request.form.get('nro_serie', '')
            equipo.area = request.form.get('area', '')
            equipo.localidad = request.form.get('localidad', '')
            equipo.responsable = request.form.get('responsable', '')
            equipo.tipo_equipo = request.form.get('tipo_equipo', '')
            equipo.subcategoria = request.form.get('subcategoria', '')
            equipo.modelo_sonda = request.form.get('modelo_sonda', '')
            equipo.modelo_equipo = request.form.get('modelo_equipo', '')
            equipo.nro_serie_sonda = request.form.get('nro_serie_sonda', '')

            # ====== ESTA ES LA LÍNEA IMPORTANTE ======
            equipo.estado = request.form.get('estado', 'Operativo')

            equipo.servicio_tecnico = request.form.get('servicio_tecnico', '')
            equipo.observaciones = nuevas_observaciones
            equipo.ultima_actualizacion = get_chile_time()

            # Nuevos campos
            equipo.nro_informe_contrastacion = request.form.get('nro_informe_contrastacion', '')
            equipo.nro_informe_mantencion = request.form.get('nro_informe_mantencion', '')
            equipo.nro_certificado_termometro = request.form.get('nro_certificado_termometro', '')

            equipo.fecha_certificado_contrastacion = datetime.strptime(request.form['fecha_certificado_contrastacion'], '%Y-%m-%d').date() if request.form.get('fecha_certificado_contrastacion') else None
            equipo.fecha_certificado_mantencion = datetime.strptime(request.form['fecha_certificado_mantencion'], '%Y-%m-%d').date() if request.form.get('fecha_certificado_mantencion') else None
            equipo.fecha_contrastacion_termometro = datetime.strptime(request.form['fecha_contrastacion_termometro'], '%Y-%m-%d').date() if request.form.get('fecha_contrastacion_termometro') else None
            equipo.fecha_despacho = datetime.strptime(request.form['fecha_despacho'], '%Y-%m-%d').date() if request.form.get('fecha_despacho') else None

            if request.form.get('fecha_contrastacion'):
                equipo.fecha_contrastacion = datetime.strptime(request.form['fecha_contrastacion'], '%Y-%m-%d').date()
            if request.form.get('fecha_certificado'):
                equipo.fecha_certificado = datetime.strptime(request.form['fecha_certificado'], '%Y-%m-%d').date()
            if request.form.get('fecha_vencimiento_insumo'):
                equipo.fecha_vencimiento_insumo = datetime.strptime(request.form['fecha_vencimiento_insumo'], '%Y-%m-%d').date()
            if request.form.get('fecha_ultima_mantencion'):
                equipo.fecha_ultima_mantencion = datetime.strptime(request.form['fecha_ultima_mantencion'], '%Y-%m-%d').date()
            if request.form.get('fecha_retorno_mantencion'):
                equipo.fecha_retorno_mantencion = datetime.strptime(request.form['fecha_retorno_mantencion'], '%Y-%m-%d').date()
            if request.form.get('fecha_envio_laboratorio'):
                equipo.fecha_envio_laboratorio = datetime.strptime(request.form['fecha_envio_laboratorio'], '%Y-%m-%d').date()

            if equipo.tipo_equipo in TIPOS_CON_VENCIMIENTO_AUTOMATICO and equipo.fecha_certificado:
                equipo.fecha_vencimiento_insumo = calcular_vencimiento_insumo(equipo.fecha_certificado)

            # Archivos
            archivos = request.files.getlist('archivos')
            for archivo in archivos:
                if archivo and archivo.filename and archivo_permitido(archivo.filename):
                    ext = archivo.filename.rsplit('.', 1)[1].lower()
                    tipo = 'foto' if ext in {'png', 'jpg', 'jpeg', 'gif'} else 'documento'
                    nombre_seguro = f"{equipo.id}_{get_chile_time().strftime('%Y%m%d_%H%M%S')}_{secure_filename(archivo.filename)}"
                    subcarpeta = 'equipos_fotos' if tipo == 'foto' else 'equipos_docs'
                    ruta = os.path.join(app.config['UPLOAD_FOLDER'], subcarpeta, nombre_seguro)
                    archivo.save(ruta)
                    if tipo == 'foto':
                        comprimir_imagen(ruta)
                    archivo_db = Archivo(
                        equipo_id=equipo.id, tipo=tipo,
                        nombre_original=archivo.filename, nombre_archivo=nombre_seguro
                    )
                    db.session.add(archivo_db)

            # Historial
            cambios = []
            historial_accion = 'CAMBIO'

            if estado_anterior != equipo.estado:
                cambios.append(f"Estado: {estado_anterior} → {equipo.estado}")
            if responsable_anterior != equipo.responsable:
                cambios.append(f"Responsable: {responsable_anterior or 'Ninguno'} → {equipo.responsable or 'Ninguno'}")

            if fecha_contrastacion_anterior != equipo.fecha_contrastacion:
                fecha_anterior = formatear_fecha(fecha_contrastacion_anterior) or 'Sin fecha'
                fecha_nueva = formatear_fecha(equipo.fecha_contrastacion) or 'Sin fecha'
                detalle = f"Contrastación: {fecha_anterior} → {fecha_nueva}"
                if equipo.nro_informe_contrastacion:
                    detalle += f" | Informe: {equipo.nro_informe_contrastacion}"
                cambios.append(detalle)
                historial_accion = 'CONTRASTACION'

            if fecha_envio_laboratorio_anterior != equipo.fecha_envio_laboratorio:
                fecha_anterior = formatear_fecha(fecha_envio_laboratorio_anterior) or 'Sin fecha'
                fecha_nueva = formatear_fecha(equipo.fecha_envio_laboratorio) or 'Sin fecha'
                cambios.append(f"Envío laboratorio: {fecha_anterior} → {fecha_nueva}")

            if fecha_ultima_mantencion_anterior != equipo.fecha_ultima_mantencion:
                fecha_anterior = formatear_fecha(fecha_ultima_mantencion_anterior) or 'Sin fecha'
                fecha_nueva = formatear_fecha(equipo.fecha_ultima_mantencion) or 'Sin fecha'
                detalle = f"Mantención: {fecha_anterior} → {fecha_nueva}"
                if equipo.nro_informe_mantencion:
                    detalle += f" | Informe: {equipo.nro_informe_mantencion}"
                cambios.append(detalle)

            if equipo.tipo_equipo in TIPOS_CON_PHMETRO:
                if modelo_sonda_anterior != equipo.modelo_sonda:
                    cambios.append(f"Modelo Sonda: {modelo_sonda_anterior or 'Ninguno'} → {equipo.modelo_sonda or 'Ninguno'}")
                if modelo_equipo_anterior != equipo.modelo_equipo:
                    cambios.append(f"Modelo Equipo: {modelo_equipo_anterior or 'Ninguno'} → {equipo.modelo_equipo or 'Ninguno'}")
                if nro_serie_sonda_anterior != equipo.nro_serie_sonda:
                    cambios.append(f"Nro Serie Sonda: {nro_serie_sonda_anterior or 'Ninguno'} → {equipo.nro_serie_sonda or 'Ninguno'}")

            if cambios:
                detalle_final = " | ".join(cambios)
                historial = Historial(
                    equipo_id=equipo.id,
                    accion=historial_accion,
                    detalle=detalle_final,
                    usuario=current_user.username,
                    responsable=equipo.responsable
                )
                db.session.add(historial)

            if nuevas_observaciones and nuevas_observaciones != observaciones_anteriores:
                historial_obs = Historial(
                    equipo_id=equipo.id,
                    accion='OBSERVACION',
                    detalle=nuevas_observaciones,
                    usuario=current_user.username,
                    responsable=equipo.responsable
                )
                db.session.add(historial_obs)

            db.session.commit()
            flash('Equipo actualizado correctamente', 'success')
            liberar_bloqueo(id)
            return redirect(url_for('dashboard'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
            liberar_bloqueo(id)

    archivos_fotos = Archivo.query.filter_by(equipo_id=equipo.id, tipo='foto').all()
    archivos_docs = Archivo.query.filter_by(equipo_id=equipo.id, tipo='documento').all()
    responsables_unicos = Responsable.query.order_by(Responsable.nombre).all()
    return render_template('equipo_form.html',
                         equipo=equipo, archivos_fotos=archivos_fotos, archivos_docs=archivos_docs,
                         LOCALIDADES=LOCALIDADES, TIPOS_EQUIPO=TIPOS_EQUIPO,
                         SUBCATEGORIAS=SUBCATEGORIAS, ESTADOS=ESTADOS, AREAS=AREAS,
                         TIPOS_CON_PHMETRO=TIPOS_CON_PHMETRO,
                         TIPOS_CON_CERTIFICADO=TIPOS_CON_CERTIFICADO,
                         TIPOS_CON_VENCIMIENTO_MANUAL=TIPOS_CON_VENCIMIENTO_MANUAL,
                         TIPOS_CON_VENCIMIENTO_AUTOMATICO=TIPOS_CON_VENCIMIENTO_AUTOMATICO,
                         TIPOS_CON_CONTRASTACION=TIPOS_CON_CONTRASTACION,
                         responsables_unicos=responsables_unicos)

@app.route('/equipo/cancelar_edicion/<int:id>')
@login_required
def cancelar_edicion(id):
    liberar_bloqueo(id)
    flash('Edición cancelada', 'info')
    return redirect(url_for('dashboard'))

@app.route('/equipo/eliminar/<int:id>')
@login_required
def eliminar_equipo(id):
    equipo = Equipo.query.get_or_404(id)
    liberar_bloqueo(id)
    for archivo in equipo.archivos:
        try:
            if archivo.tipo == 'foto':
                ruta = os.path.join(app.config['UPLOAD_FOLDER'], 'equipos_fotos', archivo.nombre_archivo)
            else:
                ruta = os.path.join(app.config['UPLOAD_FOLDER'], 'equipos_docs', archivo.nombre_archivo)
            if os.path.exists(ruta):
                os.remove(ruta)
        except:
            pass
    db.session.delete(equipo)
    db.session.commit()
    flash('Equipo eliminado correctamente', 'success')
    return redirect(url_for('dashboard'))

@app.route('/equipo/cambiar_estado/<int:id>', methods=['POST'])
@login_required
def cambiar_estado(id):
    equipo = Equipo.query.get_or_404(id)
    nuevo_estado = request.form['estado']
    nuevo_responsable = request.form.get('nuevo_responsable', '')
    servicio_tecnico = request.form.get('servicio_tecnico', '')
    fecha_contrastacion = request.form.get('fecha_contrastacion', '')
    fecha_envio_laboratorio = request.form.get('fecha_envio_laboratorio', '')
    observacion = request.form.get('observacion', '')

    estado_anterior = equipo.estado
    responsable_anterior = equipo.responsable

    hay_cambios = False
    cambios = []

    if estado_anterior != nuevo_estado:
        hay_cambios = True
        cambios.append(f"Estado: {estado_anterior} → {nuevo_estado}")

    if nuevo_responsable and nuevo_responsable != responsable_anterior:
        hay_cambios = True
        cambios.append(f"Responsable: {responsable_anterior or 'Ninguno'} → {nuevo_responsable}")

    if servicio_tecnico:
        hay_cambios = True
        cambios.append(f"Servicio Técnico: {servicio_tecnico}")

    if fecha_contrastacion:
        hay_cambios = True
        cambios.append(f"Contrastación: {fecha_contrastacion}")

    if fecha_envio_laboratorio:
        hay_cambios = True
        cambios.append(f"Envío a laboratorio: {fecha_envio_laboratorio}")

    if observacion:
        ahora = get_chile_time()
        obs_formateada = f"[{ahora.strftime('%d/%m/%Y %H:%M')}] {current_user.nombre}: {observacion}"
        if equipo.observaciones:
            equipo.observaciones = obs_formateada + "\n" + equipo.observaciones
        else:
            equipo.observaciones = obs_formateada

    if not hay_cambios and observacion:
        historial = Historial(
            equipo_id=equipo.id,
            accion='OBSERVACION',
            detalle=observacion,
            usuario=current_user.username,
            responsable=equipo.responsable
        )
        db.session.add(historial)
        db.session.commit()
        return jsonify({'success': True, 'nuevo_estado': nuevo_estado, 'mensaje': 'Observación guardada'})

    if not hay_cambios and not observacion:
        return jsonify({'success': True, 'nuevo_estado': nuevo_estado, 'mensaje': 'Sin cambios'})

    equipo.estado = nuevo_estado
    equipo.ultima_actualizacion = get_chile_time()

    if nuevo_responsable:
        equipo.responsable = nuevo_responsable
    if servicio_tecnico:
        equipo.servicio_tecnico = servicio_tecnico
    if fecha_contrastacion:
        equipo.fecha_contrastacion = datetime.strptime(fecha_contrastacion, '%Y-%m-%d').date()
    if fecha_envio_laboratorio:
        equipo.fecha_envio_laboratorio = datetime.strptime(fecha_envio_laboratorio, '%Y-%m-%d').date()

    if nuevo_estado == 'Mantencion' and estado_anterior != 'Mantencion':
        equipo.fecha_ultima_mantencion = get_chile_time().date()
    elif estado_anterior == 'Mantencion' and nuevo_estado != 'Mantencion':
        equipo.fecha_retorno_mantencion = get_chile_time().date()

    detalle_final = " | ".join(cambios)
    historial = Historial(
        equipo_id=equipo.id,
        accion='CAMBIO_ESTADO',
        detalle=detalle_final,
        usuario=current_user.username,
        responsable=equipo.responsable
    )
    db.session.add(historial)
    db.session.commit()

    return jsonify({'success': True, 'nuevo_estado': nuevo_estado})

@app.route('/equipo/duplicar/<int:id>')
@login_required
def duplicar_equipo(id):
    original = Equipo.query.get_or_404(id)
    nuevo = Equipo(
        nro_int=f"{original.nro_int}_COPY", nro_serie=original.nro_serie,
        area=original.area, localidad=original.localidad, responsable=original.responsable,
        tipo_equipo=original.tipo_equipo, subcategoria=original.subcategoria,
        modelo_sonda=original.modelo_sonda, modelo_equipo=original.modelo_equipo, nro_serie_sonda=original.nro_serie_sonda,
        estado='Operativo', observaciones=original.observaciones,
        servicio_tecnico=original.servicio_tecnico
    )
    db.session.add(nuevo)
    db.session.commit()
    historial = Historial(
        equipo_id=nuevo.id, accion='CREACION',
        detalle=f'Equipo duplicado desde {original.nro_int}',
        usuario=current_user.username, responsable=original.responsable
    )
    db.session.add(historial)
    db.session.commit()
    flash(f'Equipo duplicado: {nuevo.nro_int}', 'success')
    return redirect(url_for('dashboard'))

@app.route('/equipo/eliminar_multiples', methods=['POST'])
@login_required
def eliminar_multiples():
    ids = request.form.getlist('ids')
    for id_str in ids:
        equipo = Equipo.query.get(int(id_str))
        if equipo:
            liberar_bloqueo(int(id_str))
            for archivo in equipo.archivos:
                try:
                    if archivo.tipo == 'foto':
                        ruta = os.path.join(app.config['UPLOAD_FOLDER'], 'equipos_fotos', archivo.nombre_archivo)
                    else:
                        ruta = os.path.join(app.config['UPLOAD_FOLDER'], 'equipos_docs', archivo.nombre_archivo)
                    if os.path.exists(ruta):
                        os.remove(ruta)
                except:
                    pass
            db.session.delete(equipo)
    db.session.commit()
    flash(f'{len(ids)} equipos eliminados', 'success')
    return redirect(url_for('dashboard'))

@app.route('/equipo/historial/<int:id>')
@login_required
def ver_historial(id):
    equipo = Equipo.query.get_or_404(id)
    archivos_fotos = Archivo.query.filter_by(equipo_id=equipo.id, tipo='foto').all()
    archivos_docs = Archivo.query.filter_by(equipo_id=equipo.id, tipo='documento').all()
    responsables_unicos = Responsable.query.order_by(Responsable.nombre).all()
    return render_template('historial.html',
                         equipo=equipo,
                         archivos_fotos=archivos_fotos,
                         archivos_docs=archivos_docs,
                         formatear_fecha=formatear_fecha_hora,
                         responsables_unicos=responsables_unicos)

@app.route('/equipo/subir_archivo/<int:id>', methods=['POST'])
@login_required
def subir_archivo(id):
    equipo = Equipo.query.get_or_404(id)
    if 'archivo' not in request.files:
        flash('No se seleccionó ningún archivo', 'error')
        return redirect(url_for('ver_historial', id=equipo.id))
    archivo = request.files['archivo']
    if archivo.filename == '':
        flash('No se seleccionó ningún archivo', 'error')
        return redirect(url_for('ver_historial', id=equipo.id))
    if archivo_permitido(archivo.filename):
        ext = archivo.filename.rsplit('.', 1)[1].lower()
        tipo = 'foto' if ext in {'png', 'jpg', 'jpeg', 'gif'} else 'documento'
        if tipo == 'foto':
            fotos_existentes = Archivo.query.filter_by(equipo_id=equipo.id, tipo='foto').count()
            if fotos_existentes >= app.config['MAX_IMAGENES']:
                flash('Máximo 3 imágenes por equipo', 'error')
                return redirect(url_for('ver_historial', id=equipo.id))
        else:
            docs_existentes = Archivo.query.filter_by(equipo_id=equipo.id, tipo='documento').count()
            if docs_existentes >= app.config['MAX_DOCUMENTOS']:
                flash('Máximo 3 documentos por equipo', 'error')
                return redirect(url_for('ver_historial', id=equipo.id))
        nombre_seguro = f"{equipo.id}_{get_chile_time().strftime('%Y%m%d_%H%M%S')}_{secure_filename(archivo.filename)}"
        subcarpeta = 'equipos_fotos' if tipo == 'foto' else 'equipos_docs'
        ruta = os.path.join(app.config['UPLOAD_FOLDER'], subcarpeta, nombre_seguro)
        archivo.save(ruta)
        if tipo == 'foto':
            comprimir_imagen(ruta)
        archivo_db = Archivo(
            equipo_id=equipo.id, tipo=tipo,
            nombre_original=archivo.filename, nombre_archivo=nombre_seguro
        )
        db.session.add(archivo_db)
        db.session.commit()
        flash(f'✅ Archivo "{archivo.filename}" subido correctamente', 'success')
    else:
        flash('Tipo de archivo no permitido', 'error')
    return redirect(url_for('ver_historial', id=equipo.id))

@app.route('/equipo/eliminar_archivo/<int:archivo_id>')
@login_required
def eliminar_archivo(archivo_id):
    archivo = Archivo.query.get_or_404(archivo_id)
    equipo_id = archivo.equipo_id
    try:
        if archivo.tipo == 'foto':
            ruta = os.path.join(app.config['UPLOAD_FOLDER'], 'equipos_fotos', archivo.nombre_archivo)
        else:
            ruta = os.path.join(app.config['UPLOAD_FOLDER'], 'equipos_docs', archivo.nombre_archivo)
        if os.path.exists(ruta):
            os.remove(ruta)
    except:
        pass
    db.session.delete(archivo)
    db.session.commit()
    flash('Archivo eliminado correctamente', 'success')
    return redirect(url_for('ver_historial', id=equipo_id))

@app.route('/media/<path:filename>')
def media_files(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/exportar/excel')
@login_required
def exportar_excel():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    equipos = Equipo.query.all()
    data = []
    for e in equipos:
        alerta = obtener_alerta(e)
        data.append({
            'Nro Int': e.nro_int,
            'Nro Serie': e.nro_serie or '',
            'Área': e.area or '',
            'Localidad': e.localidad or '',
            'Responsable': e.responsable or '',
            'Tipo Equipo': e.tipo_equipo or '',
            'Subcategoría': e.subcategoria or '',
            'Modelo Sonda': e.modelo_sonda or '',
            'Modelo Equipo': e.modelo_equipo or '',
            'Nro Serie Sonda': e.nro_serie_sonda or '',
            'Estado': e.estado,
            'Servicio Técnico': e.servicio_tecnico or '',
            'Contrastación': formatear_fecha(e.fecha_contrastacion),
            'Certificado': formatear_fecha(e.fecha_certificado),
            'Vencimiento Insumo': formatear_fecha(e.fecha_vencimiento_insumo),
            'Última Mantención': formatear_fecha(e.fecha_ultima_mantencion),
            'Retorno Mantención': formatear_fecha(e.fecha_retorno_mantencion),
            'Días Inactivo': calcular_tiempo_inactividad(e),
            'Observaciones': e.observaciones or '',
            'Alerta': alerta['texto']
        })
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Inventario', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Inventario']
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        cell_font = Font(name='Segoe UI', size=10)
        header_fill = PatternFill(start_color='1E88E5', end_color='1565C0', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment_left = Alignment(horizontal='left', vertical='center')
        cell_alignment_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        even_row_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        for row_num in range(2, len(df) + 2):
            if row_num % 2 == 0:
                fill = even_row_fill
            else:
                fill = None
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = cell_font
                cell.border = thin_border
                col_name = df.columns[col_num - 1]
                if col_name in ['Nro Int', 'Días Inactivo']:
                    cell.alignment = cell_alignment_center
                else:
                    cell.alignment = cell_alignment_left
                if fill:
                    cell.fill = fill
        worksheet.row_dimensions[1].height = 25
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
        for col_num, column in enumerate(df.columns, 1):
            max_length = len(str(column))
            for row_num in range(len(df)):
                cell_value = str(df.iloc[row_num, col_num - 1])
                max_length = max(max_length, len(cell_value))
            adjusted_width = min(max_length + 2, 45)
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename=inventario_{get_chile_time().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

@app.route('/grafico')
@login_required
def grafico():
    equipos = Equipo.query.all()
    meses = {}
    for e in equipos:
        if e.tipo_equipo in TIPOS_CON_CONTRASTACION and e.fecha_contrastacion:
            fecha_venc = calcular_vencimiento_contrastacion(e.fecha_contrastacion)
            if fecha_venc:
                mes_key = fecha_venc.strftime('%Y-%m')
                meses[mes_key] = meses.get(mes_key, 0) + 1
    if not meses:
        flash('No hay datos para el gráfico', 'warning')
        return redirect(url_for('dashboard'))
    meses_ordenados = sorted(meses.keys())
    cantidades = [meses[m] for m in meses_ordenados]
    plt.figure(figsize=(10, 6))
    plt.bar(meses_ordenados, cantidades, color='#2196F3')
    plt.xlabel('Mes de Vencimiento')
    plt.ylabel('Cantidad de Equipos')
    plt.title('Equipos por Mes de Vencimiento de Contrastación')
    plt.xticks(rotation=45)
    for i, v in enumerate(cantidades):
        plt.text(i, v + 0.5, str(v), ha='center')
    img = BytesIO()
    plt.savefig(img, format='png', bbox_inches='tight')
    img.seek(0)
    plt.close()
    grafico_url = base64.b64encode(img.getvalue()).decode()
    return render_template('grafico.html', grafico_url=grafico_url)

@app.route('/enviar_correo')
@login_required
def enviar_correo():
    if not EMAIL_CONFIG.get("activado", False):
        flash('📧 El envío de correos está desactivado.', 'warning')
        return redirect(url_for('dashboard'))
    equipos = Equipo.query.all()
    vencidos = []
    proximos = []
    equipos_por_responsable = {}

    for e in equipos:
        if e.estado == 'Contrastacion':
            continue
        alerta = obtener_alerta(e)
        if 'VENCIDO' in alerta['texto'] or 'VENCIDA' in alerta['texto']:
            vencidos.append(e)
            if e.responsable:
                if e.responsable not in equipos_por_responsable:
                    equipos_por_responsable[e.responsable] = {'vencidos': [], 'proximos': []}
                equipos_por_responsable[e.responsable]['vencidos'].append(e)
        elif 'VENCE' in alerta['texto']:
            proximos.append(e)
            if e.responsable:
                if e.responsable not in equipos_por_responsable:
                    equipos_por_responsable[e.responsable] = {'vencidos': [], 'proximos': []}
                equipos_por_responsable[e.responsable]['proximos'].append(e)

    config = ConfiguracionCorreo.query.first()
    destinatarios_adicionales = []
    if config and config.destinatarios:
        try:
            destinatarios_adicionales = json.loads(config.destinatarios)
        except:
            pass

    if (vencidos or proximos) and destinatarios_adicionales:
        enviar_correo_alertas_general(vencidos, proximos, destinatarios_adicionales)

    for nombre_responsable, equipos_dict in equipos_por_responsable.items():
        responsable_db = Responsable.query.filter_by(nombre=nombre_responsable, activo=True).first()
        if responsable_db and responsable_db.email:
            enviar_correo_alertas_responsable(
                responsable_db.email,
                nombre_responsable,
                equipos_dict['vencidos'],
                equipos_dict['proximos']
            )

    flash(f'✅ Correos enviados. {len(vencidos)} vencidos, {len(proximos)} próximos.', 'success')
    return redirect(url_for('dashboard'))

@app.route('/exportar_pdf_alertas')
@login_required
def exportar_pdf_alertas():
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    import io
    equipos = Equipo.query.all()
    vencidos = []
    proximos = []
    for e in equipos:
        alerta = obtener_alerta(e)
        if 'VENCIDO' in alerta['texto'] or 'VENCIDA' in alerta['texto']:
            vencidos.append(e)
        elif 'VENCE' in alerta['texto']:
            proximos.append(e)
    if not vencidos and not proximos:
        flash('✅ No hay equipos vencidos o próximos a vencer', 'success')
        return redirect(url_for('dashboard'))
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=10*mm, leftMargin=10*mm,
                            topMargin=15*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        'Titulo', parent=styles['Heading1'], fontSize=16,
        textColor=colors.HexColor('#1E88E5'), alignment=1, spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    subtitulo_style = ParagraphStyle(
        'Subtitulo', parent=styles['Heading2'], fontSize=12,
        textColor=colors.HexColor('#c62828'), spaceAfter=8,
        fontName='Helvetica-Bold'
    )
    subtitulo_prox_style = ParagraphStyle(
        'SubtituloProx', parent=styles['Heading2'], fontSize=12,
        textColor=colors.HexColor('#f57c00'), spaceAfter=8,
        fontName='Helvetica-Bold'
    )
    normal_style = ParagraphStyle(
        'Normal', parent=styles['Normal'], fontSize=9, spaceAfter=4,
        fontName='Helvetica'
    )
    elementos = []
    elementos.append(Paragraph("SISTEMA DE INVENTARIO DE EQUIPOS", titulo_style))
    elementos.append(Paragraph(f"Reporte de Alertas - {get_chile_time().strftime('%d/%m/%Y %H:%M')}", normal_style))
    elementos.append(Spacer(1, 10))
    def crear_tabla_equipos(equipos_lista, titulo, color_titulo):
        if not equipos_lista:
            return []
        elementos_tabla = []
        elementos_tabla.append(Paragraph(titulo, subtitulo_style if color_titulo == 'rojo' else subtitulo_prox_style))
        data = [['Nro Int', 'Nro Serie', 'Área', 'Localidad', 'Responsable', 'Equipo', 'Estado', 'Alerta']]
        for e in equipos_lista:
            alerta = obtener_alerta(e)
            data.append([str(e.nro_int), e.nro_serie or '-', e.area or '-', e.localidad or '-',
                        e.responsable or '-', e.tipo_equipo or '-', e.estado, alerta['texto']])
        tabla = Table(data, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E88E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FFF5F5' if color_titulo == 'rojo' else '#FFF8E1')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D0D0')),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F8F9FA'), colors.white]),
        ]))
        elementos_tabla.append(tabla)
        elementos_tabla.append(Spacer(1, 10))
        return elementos_tabla
    elementos.extend(crear_tabla_equipos(vencidos, f"⚠️ EQUIPOS VENCIDOS ({len(vencidos)})", 'rojo'))
    elementos.extend(crear_tabla_equipos(proximos, f"⚠️ EQUIPOS PRÓXIMOS A VENCER ({len(proximos)})", 'naranja'))
    elementos.append(Paragraph("Mensaje automático del Sistema de Inventario de Equipos", normal_style))
    elementos.append(Paragraph(f"Generado el {get_chile_time().strftime('%d/%m/%Y %H:%M:%S')}", normal_style))
    doc.build(elementos)
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=alertas_vencimiento_{get_chile_time().strftime("%Y%m%d_%H%M%S")}.pdf'
    return response

@app.route('/exportar/excel/seleccionados', methods=['POST'])
@login_required
def exportar_excel_seleccionados():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    ids = request.form.getlist('ids')
    if not ids:
        flash('No hay equipos seleccionados', 'warning')
        return redirect(url_for('dashboard'))
    equipos = Equipo.query.filter(Equipo.id.in_(ids)).all()
    data = []
    for e in equipos:
        alerta = obtener_alerta(e)
        data.append({
            'Nro Int': e.nro_int,
            'Nro Serie': e.nro_serie or '',
            'Área': e.area or '',
            'Localidad': e.localidad or '',
            'Responsable': e.responsable or '',
            'Tipo Equipo': e.tipo_equipo or '',
            'Subcategoría': e.subcategoria or '',
            'Modelo Sonda': e.modelo_sonda or '',
            'Modelo Equipo': e.modelo_equipo or '',
            'Nro Serie Sonda': e.nro_serie_sonda or '',
            'Estado': e.estado,
            'Servicio Técnico': e.servicio_tecnico or '',
            'Contrastación': formatear_fecha(e.fecha_contrastacion),
            'Certificado': formatear_fecha(e.fecha_certificado),
            'Vencimiento Insumo': formatear_fecha(e.fecha_vencimiento_insumo),
            'Última Mantención': formatear_fecha(e.fecha_ultima_mantencion),
            'Retorno Mantención': formatear_fecha(e.fecha_retorno_mantencion),
            'Días Inactivo': calcular_tiempo_inactividad(e),
            'Observaciones': e.observaciones or '',
            'Alerta': alerta['texto']
        })
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Inventario', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Inventario']
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        cell_font = Font(name='Segoe UI', size=10)
        header_fill = PatternFill(start_color='1E88E5', end_color='1565C0', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment_left = Alignment(horizontal='left', vertical='center')
        cell_alignment_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        even_row_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        for row_num in range(2, len(df) + 2):
            if row_num % 2 == 0:
                fill = even_row_fill
            else:
                fill = None
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = cell_font
                cell.border = thin_border
                col_name = df.columns[col_num - 1]
                if col_name in ['Nro Int', 'Días Inactivo']:
                    cell.alignment = cell_alignment_center
                else:
                    cell.alignment = cell_alignment_left
                if fill:
                    cell.fill = fill
        worksheet.row_dimensions[1].height = 25
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
        for col_num, column in enumerate(df.columns, 1):
            max_length = len(str(column))
            for row_num in range(len(df)):
                cell_value = str(df.iloc[row_num, col_num - 1])
                max_length = max(max_length, len(cell_value))
            adjusted_width = min(max_length + 2, 45)
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename=inventario_seleccionados_{get_chile_time().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

@app.route('/admin/usuarios')
@login_required
def admin_usuarios():
    if not current_user.es_admin:
        flash('No tienes permiso para acceder a esta página', 'error')
        return redirect(url_for('dashboard'))
    usuarios = Usuario.query.order_by(Usuario.fecha_registro).all()
    return render_template('admin_usuarios.html', usuarios=usuarios)

@app.route('/admin/usuario/nuevo', methods=['GET', 'POST'])
@login_required
def admin_usuario_nuevo():
    if not current_user.es_admin:
        flash('No tienes permiso para acceder a esta página', 'error')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form['username']
        nombre = request.form['nombre']
        password = request.form['password']
        es_admin = 'es_admin' in request.form
        cambiar_password = 'cambiar_password' in request.form
        if Usuario.query.filter_by(username=username).first():
            flash('El nombre de usuario ya existe', 'error')
            return redirect(url_for('admin_usuario_nuevo'))
        nuevo = Usuario(
            username=username,
            password=hashlib.sha256(password.encode()).hexdigest(),
            nombre=nombre,
            es_admin=es_admin,
            cambiar_password=cambiar_password
        )
        db.session.add(nuevo)
        db.session.commit()
        flash(f'Usuario {username} creado correctamente', 'success')
        return redirect(url_for('admin_usuarios'))
    return render_template('admin_usuario_form.html', titulo='Nuevo Usuario')

@app.route('/admin/usuario/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def admin_usuario_editar(id):
    if not current_user.es_admin:
        flash('No tienes permiso para acceder a esta página', 'error')
        return redirect(url_for('dashboard'))
    usuario = Usuario.query.get_or_404(id)
    if request.method == 'POST':
        usuario.nombre = request.form['nombre']
        usuario.es_admin = 'es_admin' in request.form
        usuario.cambiar_password = 'cambiar_password' in request.form
        if request.form.get('password'):
            usuario.password = hashlib.sha256(request.form['password'].encode()).hexdigest()
            usuario.cambiar_password = True
        db.session.commit()
        flash(f'Usuario {usuario.username} actualizado correctamente', 'success')
        return redirect(url_for('admin_usuarios'))
    return render_template('admin_usuario_form.html', titulo='Editar Usuario', usuario=usuario)

@app.route('/admin/usuario/eliminar/<int:id>')
@login_required
def admin_usuario_eliminar(id):
    if not current_user.es_admin:
        flash('No tienes permiso para acceder a esta página', 'error')
        return redirect(url_for('dashboard'))
    usuario = Usuario.query.get_or_404(id)
    if usuario.id == current_user.id:
        flash('No puedes eliminar tu propio usuario', 'error')
        return redirect(url_for('admin_usuarios'))
    db.session.delete(usuario)
    db.session.commit()
    flash(f'Usuario {usuario.username} eliminado correctamente', 'success')
    return redirect(url_for('admin_usuarios'))

@app.route('/cambiar_password', methods=['GET', 'POST'])
@login_required
def cambiar_password():
    if request.method == 'POST':
        nueva_password = request.form['nueva_password']
        confirmar_password = request.form['confirmar_password']
        if nueva_password != confirmar_password:
            flash('Las contraseñas no coinciden', 'error')
            return redirect(url_for('cambiar_password'))
        if len(nueva_password) < 4:
            flash('La contraseña debe tener al menos 4 caracteres', 'error')
            return redirect(url_for('cambiar_password'))
        current_user.password = hashlib.sha256(nueva_password.encode()).hexdigest()
        current_user.cambiar_password = False
        db.session.commit()
        flash('Contraseña cambiada correctamente', 'success')
        return redirect(url_for('dashboard'))
    return render_template('cambiar_password.html')

@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil():
    if request.method == 'POST':
        nueva_password = request.form['nueva_password']
        confirmar_password = request.form['confirmar_password']
        if nueva_password != confirmar_password:
            flash('Las contraseñas no coinciden', 'error')
            return redirect(url_for('perfil'))
        if len(nueva_password) < 4:
            flash('La contraseña debe tener al menos 4 caracteres', 'error')
            return redirect(url_for('perfil'))
        current_user.password = hashlib.sha256(nueva_password.encode()).hexdigest()
        db.session.commit()
        flash('Contraseña cambiada correctamente', 'success')
        return redirect(url_for('dashboard'))
    return render_template('perfil.html')

@app.route('/admin/correos')
@login_required
def admin_correos():
    if not current_user.es_admin:
        flash('No tienes permiso para acceder a esta página', 'error')
        return redirect(url_for('dashboard'))
    config = ConfiguracionCorreo.query.first()
    if not config:
        config = ConfiguracionCorreo(
            frecuencia='diario',
            hora_envio='08:00',
            destinatarios=json.dumps([]),
            activo=True
        )
        db.session.add(config)
        db.session.commit()
    destinatarios_lista = json.loads(config.destinatarios) if config.destinatarios else []
    responsables = Responsable.query.order_by(Responsable.nombre).all()
    return render_template('admin_correos.html',
                         config=config,
                         destinatarios_lista=destinatarios_lista,
                         responsables=responsables,
                         EMAIL_CONFIG=EMAIL_CONFIG)

@app.route('/admin/correos/guardar', methods=['POST'])
@login_required
def admin_correos_guardar():
    if not current_user.es_admin:
        flash('No tienes permiso', 'error')
        return redirect(url_for('dashboard'))
    config = ConfiguracionCorreo.query.first()
    if not config:
        config = ConfiguracionCorreo()
        db.session.add(config)
    config.frecuencia = request.form.get('frecuencia', 'diario')
    config.hora_envio = request.form.get('hora_envio', '08:00')
    config.activo = 'activo' in request.form
    db.session.commit()
    flash('✅ Configuración de correos guardada correctamente', 'success')
    return redirect(url_for('admin_correos'))

@app.route('/admin/correos/destinatario/agregar', methods=['POST'])
@login_required
def admin_correos_agregar_destinatario():
    if not current_user.es_admin:
        flash('No tienes permiso', 'error')
        return redirect(url_for('dashboard'))
    nuevo_email = request.form.get('email', '').strip()
    if not nuevo_email:
        flash('Debe ingresar un correo electrónico', 'error')
        return redirect(url_for('admin_correos'))
    config = ConfiguracionCorreo.query.first()
    if not config:
        config = ConfiguracionCorreo()
        db.session.add(config)
    destinatarios = json.loads(config.destinatarios) if config.destinatarios else []
    if nuevo_email not in destinatarios:
        destinatarios.append(nuevo_email)
        config.destinatarios = json.dumps(destinatarios)
        db.session.commit()
        flash(f'✅ Correo {nuevo_email} agregado correctamente', 'success')
    else:
        flash('El correo ya está en la lista', 'warning')
    return redirect(url_for('admin_correos'))

@app.route('/admin/correos/destinatario/eliminar/<email>')
@login_required
def admin_correos_eliminar_destinatario(email):
    if not current_user.es_admin:
        flash('No tienes permiso', 'error')
        return redirect(url_for('dashboard'))
    config = ConfiguracionCorreo.query.first()
    if config:
        destinatarios = json.loads(config.destinatarios) if config.destinatarios else []
        if email in destinatarios:
            destinatarios.remove(email)
            config.destinatarios = json.dumps(destinatarios)
            db.session.commit()
            flash(f'✅ Correo {email} eliminado correctamente', 'success')
    return redirect(url_for('admin_correos'))

@app.route('/admin/correos/probar_envio')
@login_required
def admin_correos_probar_envio():
    if not current_user.es_admin:
        flash('No tienes permiso', 'error')
        return redirect(url_for('dashboard'))
    equipos = Equipo.query.all()
    vencidos = []
    proximos = []
    equipos_por_responsable = {}

    for e in equipos:
        if e.estado == 'Contrastacion':
            continue
        alerta = obtener_alerta(e)
        if 'VENCIDO' in alerta['texto'] or 'VENCIDA' in alerta['texto']:
            vencidos.append(e)
            if e.responsable:
                if e.responsable not in equipos_por_responsable:
                    equipos_por_responsable[e.responsable] = {'vencidos': [], 'proximos': []}
                equipos_por_responsable[e.responsable]['vencidos'].append(e)
        elif 'VENCE' in alerta['texto']:
            proximos.append(e)
            if e.responsable:
                if e.responsable not in equipos_por_responsable:
                    equipos_por_responsable[e.responsable] = {'vencidos': [], 'proximos': []}
                equipos_por_responsable[e.responsable]['proximos'].append(e)

    config = ConfiguracionCorreo.query.first()
    destinatarios_adicionales = []
    if config and config.destinatarios:
        try:
            destinatarios_adicionales = json.loads(config.destinatarios)
        except:
            pass

    if (vencidos or proximos) and destinatarios_adicionales:
        enviar_correo_alertas_general(vencidos, proximos, destinatarios_adicionales)

    for nombre_responsable, equipos_dict in equipos_por_responsable.items():
        responsable_db = Responsable.query.filter_by(nombre=nombre_responsable, activo=True).first()
        if responsable_db and responsable_db.email:
            enviar_correo_alertas_responsable(
                responsable_db.email,
                nombre_responsable,
                equipos_dict['vencidos'],
                equipos_dict['proximos']
            )

    flash('✅ Correo de prueba enviado correctamente', 'success')
    return redirect(url_for('admin_correos'))

# ========== ADMINISTRACIÓN DE RESPONSABLES ==========
@app.route('/admin/responsables/agregar', methods=['POST'])
@login_required
def admin_responsables_agregar():
    if not current_user.es_admin:
        flash('No autorizado', 'error')
        return redirect(url_for('admin_correos'))
    nombre = request.form.get('nombre', '').strip()
    email = request.form.get('email', '').strip()
    activo = 'activo' in request.form

    if not nombre:
        flash('El nombre es obligatorio', 'error')
        return redirect(url_for('admin_correos'))

    existe = Responsable.query.filter_by(nombre=nombre).first()
    if existe:
        flash(f'El responsable "{nombre}" ya existe', 'error')
        return redirect(url_for('admin_correos'))

    nuevo = Responsable(nombre=nombre, email=email if email else None, activo=activo)
    db.session.add(nuevo)
    db.session.commit()
    flash(f'✅ Responsable "{nombre}" agregado correctamente', 'success')
    return redirect(url_for('admin_correos'))

@app.route('/admin/responsables/editar/<int:id>', methods=['POST'])
@login_required
def admin_responsables_editar(id):
    if not current_user.es_admin:
        flash('No autorizado', 'error')
        return redirect(url_for('admin_correos'))
    responsable = Responsable.query.get_or_404(id)
    nombre = request.form.get('nombre', '').strip()
    email = request.form.get('email', '').strip()
    activo = 'activo' in request.form

    if not nombre:
        flash('El nombre es obligatorio', 'error')
        return redirect(url_for('admin_correos'))

    responsable.nombre = nombre
    responsable.email = email if email else None
    responsable.activo = activo
    db.session.commit()
    flash(f'✅ Responsable "{nombre}" actualizado', 'success')
    return redirect(url_for('admin_correos'))

@app.route('/admin/responsables/eliminar/<int:id>')
@login_required
def admin_responsables_eliminar(id):
    if not current_user.es_admin:
        flash('No autorizado', 'error')
        return redirect(url_for('admin_correos'))
    responsable = Responsable.query.get_or_404(id)
    nombre = responsable.nombre
    db.session.delete(responsable)
    db.session.commit()
    flash(f'✅ Responsable "{nombre}" eliminado', 'success')
    return redirect(url_for('admin_correos'))

@app.route('/api/responsables/autocomplete')
@login_required
def api_responsables_autocomplete():
    responsables = Responsable.query.filter_by(activo=True).order_by(Responsable.nombre).all()
    return jsonify([{'nombre': r.nombre, 'email': r.email} for r in responsables])

@app.route('/mantencion')
@login_required
def mantencion_dashboard():
    equipos = Equipo.query.filter(
        Equipo.estado.in_(['Mantencion', 'Fuera de Servicio'])
    ).order_by(Equipo.fecha_ultima_mantencion.desc()).all()
    for e in equipos:
        if e.estado == 'Mantencion' and e.fecha_ultima_mantencion:
            e.dias_fuera = (get_chile_time().date() - e.fecha_ultima_mantencion).days
        elif e.estado == 'Fuera de Servicio' and e.fecha_ultima_mantencion:
            e.dias_fuera = (get_chile_time().date() - e.fecha_ultima_mantencion).days
        else:
            e.dias_fuera = 0
    servicios = db.session.query(Equipo.servicio_tecnico).distinct().all()
    servicios = sorted([s[0] for s in servicios if s[0]])
    return render_template('mantencion_dashboard.html',
                         equipos=equipos,
                         servicios=servicios,
                         formatear_fecha=formatear_fecha)

@app.route('/mantencion/exportar_pdf', methods=['POST'])
@login_required
def mantencion_exportar_pdf():
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    import io

    ids = request.form.getlist('ids')
    if ids:
        equipos = Equipo.query.filter(Equipo.id.in_(ids)).all()
    else:
        equipos = Equipo.query.filter(
            Equipo.estado.in_(['Mantencion', 'Fuera de Servicio'])
        ).all()

    if not equipos:
        flash('No hay equipos para exportar', 'warning')
        return redirect(url_for('mantencion_dashboard'))

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=10*mm, leftMargin=10*mm,
                            topMargin=15*mm, bottomMargin=10*mm)

    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        'Titulo', parent=styles['Heading1'], fontSize=16,
        textColor=colors.HexColor('#1E88E5'), alignment=1, spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    normal_style = ParagraphStyle(
        'Normal', parent=styles['Normal'], fontSize=9, spaceAfter=4,
        fontName='Helvetica'
    )

    elementos = []
    elementos.append(Paragraph("SISTEMA DE INVENTARIO DE EQUIPOS", titulo_style))
    elementos.append(Paragraph(f"Reporte de Equipos en Mantención - {get_chile_time().strftime('%d/%m/%Y %H:%M')}", normal_style))
    elementos.append(Spacer(1, 10))

    data = [['Nro Int', 'Nro Serie', 'Área', 'Localidad', 'Equipo', 'Estado', 'Servicio Técnico', 'Días Fuera']]
    for e in equipos:
        if e.estado == 'Mantencion' and e.fecha_ultima_mantencion:
            dias = (get_chile_time().date() - e.fecha_ultima_mantencion).days
        else:
            dias = 0

        data.append([
            str(e.nro_int),
            e.nro_serie or '-',
            e.area or '-',
            e.localidad or '-',
            e.tipo_equipo or '-',
            e.estado,
            e.servicio_tecnico or '-',
            f'{dias} días'
        ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E88E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D0D0')),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#FFF5F5'), colors.white]),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 10))
    elementos.append(Paragraph(f"Total de equipos: {len(equipos)}", normal_style))

    doc.build(elementos)
    buffer.seek(0)

    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=mantencion_{get_chile_time().strftime("%Y%m%d_%H%M%S")}.pdf'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response

@app.route('/mantencion/exportar_excel', methods=['POST'])
@login_required
def mantencion_exportar_excel():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    ids = request.form.getlist('ids')
    if ids:
        equipos = Equipo.query.filter(Equipo.id.in_(ids)).all()
    else:
        equipos = Equipo.query.filter(
            Equipo.estado.in_(['Mantencion', 'Fuera de Servicio'])
        ).all()
    data = []
    for e in equipos:
        if e.estado == 'Mantencion' and e.fecha_ultima_mantencion:
            dias = (get_chile_time().date() - e.fecha_ultima_mantencion).days
        elif e.estado == 'Fuera de Servicio' and e.fecha_ultima_mantencion:
            dias = (get_chile_time().date() - e.fecha_ultima_mantencion).days
        else:
            dias = 0
        data.append({
            'Nro Int': e.nro_int,
            'Nro Serie': e.nro_serie or '',
            'Área': e.area or '',
            'Localidad': e.localidad or '',
            'Equipo': e.tipo_equipo or '',
            'Estado': e.estado,
            'Servicio Técnico': e.servicio_tecnico or '',
            'Fecha entrada': formatear_fecha(e.fecha_ultima_mantencion),
            'Días fuera': dias,
            'Responsable': e.responsable or '',
            'Observaciones': e.observaciones or ''
        })
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Mantención', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Mantención']
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        cell_font = Font(name='Segoe UI', size=10)
        header_fill = PatternFill(start_color='1E88E5', end_color='1565C0', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment_left = Alignment(horizontal='left', vertical='center')
        cell_alignment_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        even_row_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        for row_num in range(2, len(df) + 2):
            if row_num % 2 == 0:
                fill = even_row_fill
            else:
                fill = None
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = cell_font
                cell.border = thin_border
                col_name = df.columns[col_num - 1]
                if col_name in ['Nro Int', 'Días fuera']:
                    cell.alignment = cell_alignment_center
                else:
                    cell.alignment = cell_alignment_left
                if fill:
                    cell.fill = fill
        worksheet.row_dimensions[1].height = 25
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
        for col_num, column in enumerate(df.columns, 1):
            max_length = len(str(column))
            for row_num in range(len(df)):
                cell_value = str(df.iloc[row_num, col_num - 1])
                max_length = max(max_length, len(cell_value))
            adjusted_width = min(max_length + 2, 45)
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename=mantencion_{get_chile_time().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

@app.route('/contrastacion')
@login_required
def contrastacion_dashboard():
    equipos = Equipo.query.filter(
        Equipo.estado == 'Contrastacion'
    ).order_by(Equipo.fecha_envio_laboratorio.desc()).all()
    for e in equipos:
        if e.fecha_envio_laboratorio:
            e.dias_transcurridos = dias_habiles(e.fecha_envio_laboratorio, get_chile_time().date())
            if e.dias_transcurridos <= 10:
                e.estado_alerta = 'normal'
            elif e.dias_transcurridos <= 15:
                e.estado_alerta = 'atencion'
            else:
                e.estado_alerta = 'urgente'
        else:
            e.dias_transcurridos = 0
            e.estado_alerta = 'normal'
    return render_template('contrastacion_dashboard.html',
                         equipos=equipos,
                         formatear_fecha=formatear_fecha)

@app.route('/contrastacion/exportar_pdf', methods=['POST'])
@login_required
def contrastacion_exportar_pdf():
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    import io

    ids = request.form.getlist('ids')
    if ids:
        equipos = Equipo.query.filter(Equipo.id.in_(ids)).all()
    else:
        equipos = Equipo.query.filter(Equipo.estado == 'Contrastacion').all()

    if not equipos:
        flash('No hay equipos para exportar', 'warning')
        return redirect(url_for('contrastacion_dashboard'))

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=10*mm, leftMargin=10*mm,
                            topMargin=15*mm, bottomMargin=10*mm)

    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        'Titulo', parent=styles['Heading1'], fontSize=16,
        textColor=colors.HexColor('#1E88E5'), alignment=1, spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    normal_style = ParagraphStyle(
        'Normal', parent=styles['Normal'], fontSize=9, spaceAfter=4,
        fontName='Helvetica'
    )

    elementos = []
    elementos.append(Paragraph("SISTEMA DE INVENTARIO DE EQUIPOS", titulo_style))
    elementos.append(Paragraph(f"Reporte de Equipos en Contradicción - {get_chile_time().strftime('%d/%m/%Y %H:%M')}", normal_style))
    elementos.append(Spacer(1, 10))

    data = [['Nro Int', 'Nro Serie', 'Área', 'Localidad', 'Equipo', 'Fecha Envío', 'Días en Lab', 'Estado']]
    for e in equipos:
        if e.fecha_envio_laboratorio:
            dias = (get_chile_time().date() - e.fecha_envio_laboratorio).days
            if dias <= 10:
                estado_texto = '🟢 En plazo'
            elif dias <= 14:
                estado_texto = '🟡 Atención'
            else:
                estado_texto = '🔴 Vencido'
        else:
            dias = 0
            estado_texto = '⚪ Sin fecha'

        data.append([
            str(e.nro_int),
            e.nro_serie or '-',
            e.area or '-',
            e.localidad or '-',
            e.tipo_equipo or '-',
            formatear_fecha(e.fecha_envio_laboratorio) or '-',
            f'{dias} días',
            estado_texto
        ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E88E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D0D0')),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#FFF8E1'), colors.white]),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 10))
    elementos.append(Paragraph(f"Total de equipos: {len(equipos)}", normal_style))

    doc.build(elementos)
    buffer.seek(0)

    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=contrastacion_{get_chile_time().strftime("%Y%m%d_%H%M%S")}.pdf'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response

@app.route('/contrastacion/exportar_excel', methods=['POST'])
@login_required
def contrastacion_exportar_excel():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    ids = request.form.getlist('ids')
    if ids:
        equipos = Equipo.query.filter(Equipo.id.in_(ids)).all()
    else:
        equipos = Equipo.query.filter(Equipo.estado == 'Contrastacion').all()
    data = []
    for e in equipos:
        if e.fecha_envio_laboratorio:
            dias = dias_habiles(e.fecha_envio_laboratorio, get_chile_time().date())
            if dias <= 10:
                estado_plazo = '🟢 En plazo'
            elif dias <= 15:
                estado_plazo = '🟡 Atención'
            else:
                estado_plazo = '🔴 Vencido'
        else:
            dias = 0
            estado_plazo = '⚪ Sin fecha'
        data.append({
            'Nro Int': e.nro_int,
            'Nro Serie': e.nro_serie or '',
            'Área': e.area or '',
            'Localidad': e.localidad or '',
            'Equipo': e.tipo_equipo or '',
            'Fecha Envío': formatear_fecha(e.fecha_envio_laboratorio),
            'Días en Lab': dias,
            'Estado Plazo': estado_plazo,
            'Responsable': e.responsable or '',
            'Observaciones': e.observaciones or ''
        })
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Contrastación', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Contrastación']
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        cell_font = Font(name='Segoe UI', size=10)
        header_fill = PatternFill(start_color='1E88E5', end_color='1565C0', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment_left = Alignment(horizontal='left', vertical='center')
        cell_alignment_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        even_row_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        for row_num in range(2, len(df) + 2):
            if row_num % 2 == 0:
                fill = even_row_fill
            else:
                fill = None
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = cell_font
                cell.border = thin_border
                col_name = df.columns[col_num - 1]
                if col_name in ['Nro Int', 'Días en Lab']:
                    cell.alignment = cell_alignment_center
                else:
                    cell.alignment = cell_alignment_left
                if fill:
                    cell.fill = fill
        worksheet.row_dimensions[1].height = 25
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
        for col_num, column in enumerate(df.columns, 1):
            max_length = len(str(column))
            for row_num in range(len(df)):
                cell_value = str(df.iloc[row_num, col_num - 1])
                max_length = max(max_length, len(cell_value))
            adjusted_width = min(max_length + 2, 45)
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename=contrastacion_{get_chile_time().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

@app.route('/equipo/agregar_observacion', methods=['POST'])
@login_required
def agregar_observacion():
    data = request.get_json()
    equipo = Equipo.query.get_or_404(data['equipo_id'])
    observacion = data['observacion'].strip()

    if not observacion:
        return jsonify({'success': False, 'error': 'La observación no puede estar vacía'})

    ahora = get_chile_time()
    observacion_formateada = f"[{ahora.strftime('%d/%m/%Y %H:%M')}] {current_user.nombre}: {observacion}"
    if equipo.observaciones:
        equipo.observaciones = observacion_formateada + "\n" + equipo.observaciones
    else:
        equipo.observaciones = observacion_formateada

    from datetime import timedelta
    hace_5_segundos = get_chile_time() - timedelta(seconds=5)
    historial_existente = Historial.query.filter(
        Historial.equipo_id == equipo.id,
        Historial.accion == 'CAMBIO_ESTADO',
        Historial.detalle.contains(f"Observación: {observacion}"),
        Historial.fecha_hora >= hace_5_segundos
    ).first()

    if historial_existente:
        db.session.commit()
        return jsonify({'success': True, 'mensaje': 'Observación ya guardada en cambio de estado'})

    historial = Historial(
        equipo_id=equipo.id,
        accion='OBSERVACION',
        detalle=observacion,
        usuario=current_user.username,
        responsable=equipo.responsable
    )
    db.session.add(historial)
    db.session.commit()

    return jsonify({'success': True})

def verificar_y_crear_respaldo():
    backup_dir = os.path.join(BASE_DIR, 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    chile_tz = pytz.timezone('America/Santiago')
    ahora = datetime.now(chile_tz)
    respaldos = []
    for f in os.listdir(backup_dir):
        if f.startswith('inventario_') and f.endswith('.db'):
            ruta = os.path.join(backup_dir, f)
            mtime = datetime.fromtimestamp(os.path.getmtime(ruta), tz=chile_tz)
            respaldos.append((ruta, mtime))
    if respaldos:
        respaldos.sort(key=lambda x: x[1], reverse=True)
        ultimo_respaldo = respaldos[0][1]
        horas_desde_ultimo = (ahora - ultimo_respaldo).total_seconds() / 3600
        if horas_desde_ultimo < 24:
            return
    fecha = ahora.strftime('%Y%m%d_%H%M%S')
    backup_file = os.path.join(backup_dir, f'inventario_auto_{fecha}.db')
    db_path = os.path.join(BASE_DIR, 'inventario.db')
    shutil.copy2(db_path, backup_file)
    try:
        archivos = []
        for f in os.listdir(backup_dir):
            if f.startswith('inventario_') and f.endswith('.db'):
                ruta = os.path.join(backup_dir, f)
                archivos.append((ruta, os.path.getmtime(ruta)))
        archivos.sort(key=lambda x: x[1], reverse=True)
        for ruta, _ in archivos[30:]:
            os.remove(ruta)
    except:
        pass

def crear_respaldo(tipo='auto'):
    try:
        backup_dir = os.path.join(BASE_DIR, 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        fecha = get_chile_time().strftime('%Y%m%d_%H%M%S')
        backup_file = os.path.join(backup_dir, f'inventario_{tipo}_{fecha}.db')
        db_path = os.path.join(BASE_DIR, 'inventario.db')
        shutil.copy2(db_path, backup_file)
        limpiar_respaldos_antiguos(backup_dir, 30)
        return backup_file
    except Exception as e:
        print(f"Error al crear respaldo: {e}")
        return None

def limpiar_respaldos_antiguos(directorio, mantener=30):
    try:
        archivos = []
        for f in os.listdir(directorio):
            if f.startswith('inventario_') and f.endswith('.db'):
                ruta = os.path.join(directorio, f)
                archivos.append((ruta, os.path.getmtime(ruta)))
        archivos.sort(key=lambda x: x[1], reverse=True)
        for ruta, _ in archivos[mantener:]:
            os.remove(ruta)
    except:
        pass

@app.route('/backups')
@login_required
def listar_backups():
    if not current_user.es_admin:
        flash('No tienes permiso para acceder a esta página', 'error')
        return redirect(url_for('dashboard'))
    backup_dir = os.path.join(BASE_DIR, 'backups')
    backups = []
    if os.path.exists(backup_dir):
        for f in os.listdir(backup_dir):
            if f.startswith('inventario_') and f.endswith('.db'):
                ruta = os.path.join(backup_dir, f)
                stats = os.stat(ruta)
                partes = f.replace('inventario_', '').replace('.db', '').split('_')
                tipo = partes[0] if len(partes) > 0 else 'auto'
                backups.append({
                    'nombre': f,
                    'tipo': tipo,
                    'fecha': datetime.fromtimestamp(stats.st_mtime).strftime('%d/%m/%Y %H:%M:%S'),
                    'tamaño': f"{stats.st_size / 1024:.2f} KB",
                    'ruta': ruta
                })
        backups.sort(key=lambda x: x['fecha'], reverse=True)
    return render_template('backups.html', backups=backups)

@app.route('/backup/crear')
@login_required
def backup_crear():
    if not current_user.es_admin:
        flash('No tienes permiso', 'error')
        return redirect(url_for('dashboard'))
    archivo = crear_respaldo('manual')
    if archivo:
        flash(f'✅ Respaldo creado correctamente', 'success')
    else:
        flash('❌ Error al crear el respaldo', 'error')
    return redirect(url_for('listar_backups'))

@app.route('/backup/restaurar/<nombre>', methods=['POST'])
@login_required
def backup_restaurar(nombre):
    if not current_user.es_admin:
        flash('No tienes permiso', 'error')
        return redirect(url_for('dashboard'))
    confirmacion = request.form.get('confirmacion', '')
    if confirmacion != 'RESTAURAR':
        flash('Debes escribir "RESTAURAR" para confirmar', 'error')
        return redirect(url_for('listar_backups'))
    try:
        backup_dir = os.path.join(BASE_DIR, 'backups')
        backup_path = os.path.join(backup_dir, nombre)
        db_path = os.path.join(BASE_DIR, 'inventario.db')
        if not os.path.exists(backup_path):
            flash('El archivo de respaldo no existe', 'error')
            return redirect(url_for('listar_backups'))
        crear_respaldo('emergencia')
        shutil.copy2(backup_path, db_path)
        historial = Historial(
            equipo_id=1,
            accion='RESTAURACION',
            detalle=f'Base de datos restaurada desde respaldo: {nombre}',
            usuario=current_user.username,
            responsable=current_user.nombre
        )
        db.session.add(historial)
        db.session.commit()
        flash(f'✅ Base de datos restaurada correctamente desde {nombre}', 'success')
        logout_user()
        return redirect(url_for('login'))
    except Exception as e:
        flash(f'❌ Error al restaurar: {str(e)}', 'error')
        return redirect(url_for('listar_backups'))

@app.route('/backup/eliminar/<nombre>')
@login_required
def backup_eliminar(nombre):
    if not current_user.es_admin:
        flash('No tienes permiso', 'error')
        return redirect(url_for('dashboard'))
    try:
        backup_dir = os.path.join(BASE_DIR, 'backups')
        backup_path = os.path.join(backup_dir, nombre)
        if os.path.exists(backup_path):
            os.remove(backup_path)
            flash(f'✅ Respaldo {nombre} eliminado', 'success')
        else:
            flash('El archivo no existe', 'error')
    except Exception as e:
        flash(f'Error al eliminar: {str(e)}', 'error')
    return redirect(url_for('listar_backups'))

def verificar_y_enviar_alertas():
    with app.app_context():
        config = ConfiguracionCorreo.query.first()
        if not config or not config.activo:
            return
        ahora = get_chile_time()
        if config.ultimo_envio and config.ultimo_envio.date() == ahora.date():
            return
        hora_config = config.hora_envio.split(':')
        hora_actual = ahora.hour
        minuto_actual = ahora.minute
        if hora_actual < int(hora_config[0]):
            return
        if hora_actual == int(hora_config[0]) and minuto_actual < int(hora_config[1]):
            return
        if config.frecuencia == 'diario':
            enviar_alertas_automatico()
        elif config.frecuencia == 'semanal':
            if ahora.weekday() == 0:
                enviar_alertas_automatico()
        elif config.frecuencia == 'mensual':
            if ahora.day == 1:
                enviar_alertas_automatico()

def enviar_alertas_automatico():
    equipos = Equipo.query.all()
    vencidos = []
    proximos = []
    equipos_por_responsable = {}

    for e in equipos:
        if e.estado == 'Contrastacion':
            continue
        alerta = obtener_alerta(e)
        if 'VENCIDO' in alerta['texto'] or 'VENCIDA' in alerta['texto']:
            vencidos.append(e)
            if e.responsable:
                if e.responsable not in equipos_por_responsable:
                    equipos_por_responsable[e.responsable] = {'vencidos': [], 'proximos': []}
                equipos_por_responsable[e.responsable]['vencidos'].append(e)
        elif 'VENCE' in alerta['texto']:
            proximos.append(e)
            if e.responsable:
                if e.responsable not in equipos_por_responsable:
                    equipos_por_responsable[e.responsable] = {'vencidos': [], 'proximos': []}
                equipos_por_responsable[e.responsable]['proximos'].append(e)

    config = ConfiguracionCorreo.query.first()
    destinatarios_adicionales = []
    if config and config.destinatarios:
        try:
            destinatarios_adicionales = json.loads(config.destinatarios)
        except:
            pass

    if (vencidos or proximos) and destinatarios_adicionales:
        enviar_correo_alertas_general(vencidos, proximos, destinatarios_adicionales)

    for nombre_responsable, equipos_dict in equipos_por_responsable.items():
        responsable_db = Responsable.query.filter_by(nombre=nombre_responsable, activo=True).first()
        if responsable_db and responsable_db.email:
            enviar_correo_alertas_responsable(
                responsable_db.email,
                nombre_responsable,
                equipos_dict['vencidos'],
                equipos_dict['proximos']
            )

    if config:
        config.ultimo_envio = get_chile_time()
        db.session.commit()

@app.route('/api/verificar_bloqueos')
@login_required
def api_verificar_bloqueos():
    equipos = Equipo.query.all()
    resultado = {}
    for equipo in equipos:
        bloqueado, usuario, fecha = verificar_bloqueo(equipo.id)
        resultado[equipo.id] = {
            'bloqueado': bloqueado,
            'usuario': usuario if bloqueado else None
        }
    return jsonify(resultado)

@app.route('/api/puede_editar/<int:equipo_id>')
@login_required
def api_puede_editar(equipo_id):
    bloqueado, usuario, fecha = verificar_bloqueo(equipo_id)
    if bloqueado and usuario != current_user.nombre:
        return jsonify({'puede': False, 'usuario': usuario})
    return jsonify({'puede': True})

@app.route('/api/verificar_nro_int')
@login_required
def api_verificar_nro_int():
    nro_int = request.args.get('nro_int', '')
    if not nro_int:
        return jsonify({'existe': False})
    existe = Equipo.query.filter(Equipo.nro_int == nro_int).first() is not None
    return jsonify({'existe': existe})  # <--- CORRECTO: DEVUELVE LA VARIABLE

@app.route('/api/verificar_nro_serie')
@login_required
def api_verificar_nro_serie():
    nro_serie = request.args.get('nro_serie', '')
    if not nro_serie:
        return jsonify({'existe': False})
    existe = Equipo.query.filter(Equipo.nro_serie == nro_serie).first() is not None
    return jsonify({'existe': existe})  # <--- CORRECTO: DEVUELVE LA VARIABLE

@app.route('/limpiar_flash', methods=['POST'])
def limpiar_flash():
    session.pop('_flashes', None)
    return '', 200

with app.app_context():
    db.create_all()
    if not Usuario.query.filter_by(username='admin').first():
        admin = Usuario(
            username='admin', password=hashlib.sha256('admin123'.encode()).hexdigest(),
            nombre='Administrador', es_admin=True, cambiar_password=False
        )
        db.session.add(admin)
        db.session.commit()
        print("Usuario admin creado: admin / admin123")
    if not ConfiguracionCorreo.query.first():
        config = ConfiguracionCorreo(
            frecuencia='diario',
            hora_envio='08:00',
            destinatarios=json.dumps([]),
            activo=True
        )
        db.session.add(config)
        db.session.commit()
        print("Configuración de correo creada")

@app.route('/cron/enviar_alertas')
def cron_enviar_alertas():
    clave = request.args.get('clave', '')
    if clave != 'suralis2024':
        return 'No autorizado', 401
    with app.app_context():
        enviar_alertas_automatico()
        return 'Alertas enviadas correctamente', 200

if __name__ == '__main__':
    app.run(debug=True)